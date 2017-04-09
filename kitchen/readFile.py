# -*- coding: utf-8 -*-
"""
Created on Sat Apr  8 19:45:45 2017

@author: saulo
"""

#!/usr/bin/env python
# -*- coding: utf-8 -*-

#pip install openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

#pip install mysql-connector-python-rf
import mysql.connector


import codecs

def podeSerInteiro(dado):
    try: 
        int(dado)
        return True
    except ValueError:
        return False

def proximaColuna(colunaAtual):
    listaDeLetras="ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if colunaAtual[-1] == "Z":        
        colunaAtual=(len(colunaAtual)+1)*"A"
        return colunaAtual
    else:
        indice = listaDeLetras.find(colunaAtual[-1])
        novaLetra = listaDeLetras[indice+1]
        colunaAtual=colunaAtual[:-1]
        return colunaAtual+novaLetra

def executarProcedure(nomeProcedure,linha,mySqlConfig):
    print nomeProcedure,'\n',linha,'\n',mySqlConfig
    conection = mysql.connector.connect(**mySqlConfig)
    cursor = conection.cursor()
    cursor.callproc(operation, linha)
    conection.close()
        
            
def abrirArquivo(path,mySqlConfig, nomeProcedure):
    planilha = load_workbook(filename = path)
        
    # nome das abas sheetnames
    for aba in planilha.sheetnames:
        count=1
        sheet_ranges=planilha[aba]
        coluna="A"
        colunas=[]
        dicionarioDeColunas={}
        while unicode(sheet_ranges[coluna+str(count)].value).encode('utf-8')!=unicode("None").encode('utf-8'):
            colunas.append(coluna)
            dicionarioDeColunas[unicode(sheet_ranges[coluna+str(count)].value)]=coluna
            print sheet_ranges[coluna+str(count)].value
            coluna = coluna[:-1]+proximaColuna(coluna[-1])
        coluna="A"
        
        
        data = []
        while unicode(sheet_ranges[coluna+str(count)].value).encode('utf-8')!=unicode("None").encode('utf-8'):
            linha=[]            
            for coluna in colunas:
                dado = sheet_ranges[coluna+str(count)].value
                
                if type(dado) in [long,unicode]:
                    if podeSerInteiro(dado):
                        linha.append(int(dado))
                        
                    elif unicode(":").encode('utf-8') in unicode(dado).encode('utf-8'): 
                        dado=unicode(dado).encode('utf-8')
                        dado=dado[:2]+"\\"+dado[2:6]+"\\"+dado[6:]
                        linha.append(unicode(dado).encode('utf-8'))
                        
                    else:
                        linha.append(unicode(dado).encode('utf-8'))
                    
                elif unicode(dado).encode('utf-8')==unicode("None").encode('utf-8'):
                    linha.append(unicode("--NA--").encode('utf-8'))
                
                elif type(dado) not in [long,unicode]:
                    linha.append(dado.strftime("%y//%m//%d"))
                    
                    
                else:
                    linha.append(dado)
                    
                coluna = coluna[:-1]+proximaColuna(coluna[-1])
            executarProcedure(nomeProcedure,linha,mySqlConfig)    
            count+=1

            coluna="A"
        return data
        
mySqlConfig = {
  'user': 'saude',
  'password': 'asd123asd',
  'host': 'juniper.arvixe.com'
}

#abrirArquivo('BRASÍLIA - RELATÓRIO PRIMEIRO DESTINO DE ENCAMINHAMENTO.xlsx')
abrirArquivo('BRASÍLIA- Banco de Dados SMS SÃO PAULO E REDE.xlsx',mySqlConfig,"nomeProcedure")

